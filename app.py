from flask import Flask, render_template, request, redirect, url_for, session, flash
from flask_mysqldb import MySQL
from werkzeug.security import generate_password_hash, check_password_hash
from functools import wraps
from werkzeug.utils import secure_filename
import os
import MySQLdb.cursors
from decimal import Decimal
from xhtml2pdf import pisa
from flask import make_response,send_file,current_app
from io import BytesIO
from openpyxl import Workbook
from collections import defaultdict
import uuid
from pathlib import Path
from flask import make_response
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import cm
from reportlab.platypus import Table, TableStyle, Image, Paragraph, SimpleDocTemplate, Spacer
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib import colors
from babel.numbers import format_currency
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta


app = Flask(__name__, instance_relative_config=True)
app.config.from_pyfile('config.py')
from instance.db import mysql  # Import mysql dari db.py
app.secret_key = app.config['SECRET_KEY']

mysql.init_app(app)  # pastikan inisialisasi

persen = 0.01

def format_rupiah_excel(angka):
    # Mengubah angka menjadi format seperti 100.000
    return "{:,.0f}".format(angka).replace(",", ".")

# def format_rupiah(angka):
#     try:
#         return "{:,.0f}".format(float(angka)).replace(",", ".")
#     except:
#         return "0"

def format_rupiah(angka):
    return format_currency(angka, 'IDR', locale='id_ID', format=u'¤#,##0')

def generate_nomor_pembayaran():
    return uuid.uuid4().hex[:12].upper()

def get_absolute_foto_path(filename: str) -> str:
    return Path("static/uploads") / filename

@app.route('/')
def index():
    return redirect(url_for('dashboard'))

# ======= LOGIN REQUIRED DECORATOR =======
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated_function

def admin_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if session.get('role') != 'admin':
            flash('Akses khusus admin!', 'danger')
            return redirect(url_for('dashboard'))
        return f(*args, **kwargs)
    return decorated_function

# ========== LOGIN ==========
@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        email = request.form['email']
        password_input = request.form['password']

        cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        cur.execute("SELECT * FROM users WHERE email = %s", (email,))
        user = cur.fetchone()
        cur.close()

        if user:
            if check_password_hash(user['password'], password_input):
                session['user_id'] = user['id']           # <= PENTING
                session['name'] = user['name']             # <= PENTING
                session['role'] = user['role']             # <= PENTING
                session['email'] = user['email']
                flash('Login berhasil!', 'success')
                return redirect(url_for('dashboard'))
            else:
                flash('Password salah', 'danger')
        else:
            flash('Email atau password salah!', 'danger')
    return render_template('login.html')

@app.route('/profile')
@login_required
def profile():
    # Ambil user_id dan profile_id dari session
    user_id = session.get('user_id')

    if not user_id:
        flash('Anda perlu login terlebih dahulu.', 'danger')
        return redirect(url_for('login'))
    
    # Mengambil data user berdasarkan user_id
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    cur.execute("SELECT * FROM users WHERE id = %s", (user_id,))
    user = cur.fetchone()

    # Render template profile dengan data user dan profile
    return render_template('profile.html', user=user)

# ========== DASHBOARD ==========
@app.route('/dashboard')
@login_required
def dashboard():
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    
    # Ambil data total pembelian per bulan (untuk grafik admin)
    cur.execute("""
        SELECT DATE_FORMAT(tgl_pembelian, '%Y-%m') AS bulan, COUNT(*) AS jumlah_pembelian
        FROM pembelian
        GROUP BY bulan
        ORDER BY bulan
    """)
    hasil = cur.fetchall()
    bulan_list = [row['bulan'] for row in hasil]
    jumlah_pembelian = [row['jumlah_pembelian'] for row in hasil]

    bulan = datetime.now().month
    tahun = datetime.now().year

    # Hitung total pembelian bulan ini untuk user yang login
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    cur.execute("""
        SELECT COUNT(*) AS total FROM pembelian 
        WHERE MONTH(tgl_pembelian) = %s AND YEAR(tgl_pembelian) = %s AND user_id = %s
    """, (bulan, tahun, session['user_id']))

    result = cur.fetchone()
    total_pembelian_bulan_ini = result['total'] if result else 0

    cur.close()

    return render_template("dashboard.html",
        bulan_list=bulan_list,
        jumlah_pembelian=jumlah_pembelian,
        total_pembelian_bulan_ini=total_pembelian_bulan_ini
    )

@app.route('/users')
@login_required
@admin_required
def list_users():
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    # Join users dan profile
    cur.execute("""
        SELECT * FROM users
    """)
    users = cur.fetchall()
    cur.close()
    return render_template('list_users.html', users=users)

@app.route('/user/export-excel')
@login_required
def export_users_excel():
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    # Ambil data yang sesuai
    cur.execute("""
        SELECT name, email, role
        FROM users
    """)
    users_list = cur.fetchall()

    # Buat workbook Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Daftar Users"

    # Header kolom
    headers = ['Nama User', 'Email', 'Role']
    ws.append(headers)

    # Gaya header
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F81BD")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    # Isi data
    for row_index, user in enumerate(users_list, start=2):
        ws.cell(row=row_index, column=1, value=user['name'])
        ws.cell(row=row_index, column=2, value=user['email'])
        ws.cell(row=row_index, column=3, value=user['role'])
        for col in range(1, 4):
            ws.cell(row=row_index, column=col).border = border
            ws.cell(row=row_index, column=col).alignment = Alignment(vertical="center")

    # Menyesuaikan lebar kolom otomatis
    for column_cells in ws.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = length + 2

    # Simpan ke dalam memori
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return send_file(
        file_stream,
        as_attachment=True,
        download_name='daftar_user.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/user/tambah', methods=['GET', 'POST'])
@login_required
@admin_required
def tambah_user():
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    if request.method == 'POST':
        name = request.form['name']
        email = request.form['email']
        password = generate_password_hash(request.form['password'])
        role = request.form['role']
        
        # Validasi: Non-member tidak boleh isi no_hp
        # if status != 'member' and no_hp.strip():
        #     flash('Non-member tidak diperbolehkan memiliki No HP.', 'danger')
        #     return redirect(request.url)

        # Insert ke tabel users
        cur.execute("""
            INSERT INTO users (name, email, password, role)
            VALUES (%s, %s, %s, %s)
        """, (name, email, password, role))

        mysql.connection.commit()
        cur.close()

        flash('User dan Profile berhasil ditambahkan', 'success')
        return redirect(url_for('list_users'))

    return render_template('tambah_user.html')

@app.route('/user/edit/<int:user_id>', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_user(user_id):
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    if request.method == 'POST':
        name = request.form['name']
        email = request.form['email']
        role = request.form['role']

        # cur.execute("SELECT status FROM profile WHERE user_id = %s", (user_id,))
        # current_status = cur.fetchone()['status']

        # Update users
        cur.execute("""
            UPDATE users SET name=%s, email=%s, role=%s
            WHERE id=%s
        """, (name, email, role, user_id))

        # # Ambil profile id untuk user
        # cur.execute("SELECT id FROM profile WHERE user_id = %s", (user_id,))
        # profile = cur.fetchone()

        # # Jika status berubah dari member ke non_member, set point ke 0
        # if current_status == 'member' and status == 'non_member':
        #     cur.execute("UPDATE profile SET point = 0 WHERE user_id = %s", (user_id,))
        # else:
        #     # Jika tidak berubah ke non_member, kita biarkan atau update sesuai inputan point
        #     if point is not None:
        #         cur.execute("UPDATE profile SET point = %s WHERE user_id = %s", (point, user_id))

        # # Jika status member, update join_date
        # join_date = datetime.now() if status.lower() == 'member' else None

        # # Proses no_hp
        # no_hp = request.form['no_hp'] if status == 'member' else None

        # cur.execute("""
        #     UPDATE profile SET status=%s, join_date=%s, no_hp=%s WHERE user_id=%s
        # """, (status, join_date, no_hp, user_id))

        mysql.connection.commit()
        cur.close()
        flash('User berhasil diperbarui', 'success')
        return redirect(url_for('list_users'))

    # GET data user
    cur.execute("""
        SELECT * FROM users
        WHERE id = %s
    """, (user_id,))
    user = cur.fetchone()
    cur.close()

    return render_template('edit_user.html', user=user)

@app.route('/users/delete/<int:user_id>', methods=['POST'])
@login_required
@admin_required
def delete_user(user_id):
    cur = mysql.connection.cursor()
    logged_in_user_id = session.get('user_id')  # Mendapatkan user_id yang sedang login
    print(f"User ID yang login: {logged_in_user_id}")
    
    # Cek apakah user yang akan dihapus memiliki riwayat pembelian
    cur.execute("SELECT COUNT(*) FROM pembelian WHERE user_id = %s", (user_id,))
    count = cur.fetchone()[0]
    
    # Pastikan data user yang akan dihapus ada
    cur.execute("SELECT id FROM users WHERE id = %s", (user_id,))
    user = cur.fetchone()
    print(f"User ID yang akan dihapus: {user}")

    # Pastikan user lain tetap bisa dihapus
    if not user:
        flash('User tidak ditemukan.', 'danger')
        return redirect(url_for('list_users'))

    # Mencegah penghapusan akun user yang sedang login
    if user[0] == logged_in_user_id:
        flash('Tidak bisa menghapus akun Anda sendiri.', 'danger')
        return redirect(url_for('list_users'))

    # Mencegah penghapusan user dengan riwayat pembelian
    if count > 0:
        flash('Tidak bisa menghapus user karena memiliki riwayat pembelian.', 'danger')
        return redirect(url_for('list_users'))

    # Jika semua pengecekan lolos, lanjutkan penghapusan
    cur.execute("DELETE FROM users WHERE id = %s", (user_id,))
    mysql.connection.commit()
    cur.close()
    flash('User berhasil dihapus.', 'success')
    return redirect(url_for('list_users'))


# ========== KONFIGURASI FOTO ==========
UPLOAD_FOLDER = 'static/uploads'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg'}

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# ========== PRODUK LIST ==========
@app.route('/produk')
@login_required
def produk_list():
    q = request.args.get('q', '')  # Ambil query dari parameter URL
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    if q:
        cur.execute("SELECT * FROM produk WHERE nama_produk LIKE %s", ('%' + q + '%',))
    else:
        cur.execute("SELECT * FROM produk")
    produk = cur.fetchall()
    cur.close()
    return render_template('produk_list.html', produk=produk)

@app.route('/produk/export-excel')
@login_required
def export_produk_excel():
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    q = request.args.get('q', '').strip()

    query = """
        SELECT nama_produk, harga, stok
        FROM produk
        WHERE 1=1
    """
    params = []

    if q:
        query += " AND nama_produk LIKE %s"
        params.append(f"%{q}%")

    cur.execute(query, tuple(params))
    produk_list = cur.fetchall()
    cur.close()

    # Buat file Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Daftar Produk"

    # Header
    headers = ['Nama Produk', 'Harga', 'Stok']
    ws.append(headers)

    # Gaya untuk header
    header_font = Font(bold=True, color="000000")
    header_fill = PatternFill(start_color="DDDDDD", end_color="DDDDDD", fill_type="solid")
    border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border

    # Data baris
    for row_num, p in enumerate(produk_list, start=2):
        row_data = [p['nama_produk'], format_rupiah(p['harga']), p['stok']]
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            cell.border = border

    # Auto-width kolom
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = max_length + 2
        ws.column_dimensions[column].width = adjusted_width

    # Simpan ke memori
    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return send_file(
        file_stream,
        as_attachment=True,
        download_name='daftar_produk.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

# ========== TAMBAH PRODUK ==========
@app.route('/produk/tambah', methods=['GET', 'POST'])
@login_required
@admin_required
def tambah_produk():
    if request.method == 'POST':
        nama = request.form['nama']
        harga = request.form['harga']
        stok = request.form['stok']
        foto = request.files['foto']

        foto_filename = None
        if foto and foto.filename != '':
            foto_filename = secure_filename(foto.filename)
            foto.save(os.path.join(app.config['UPLOAD_FOLDER'], foto_filename))

        cur = mysql.connection.cursor()
        cur.execute(
            'INSERT INTO produk (nama_produk, harga, stok, foto) VALUES (%s, %s, %s, %s)',
            (nama, harga, stok, foto_filename)
        )
        mysql.connection.commit()
        cur.close()
        return redirect(url_for('produk_list'))
    return render_template('produk_form.html')

# ========== EDIT PRODUK ==========
@app.route('/produk/edit/<int:id>', methods=['GET', 'POST'])
@login_required
@admin_required
def edit_produk(id):
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    # Ambil data produk berdasarkan id
    cur.execute("SELECT * FROM produk WHERE id = %s", (id,))
    produk = cur.fetchone()

    if request.method == 'POST':
        foto = request.files['foto']
        if foto and allowed_file(foto.filename):
            filename = secure_filename(foto.filename)
            foto.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            cur.execute("UPDATE produk SET foto=%s WHERE id=%s", (filename, id))
        nama_produk = request.form['nama_produk']
        harga = request.form['harga']
        stok = request.form['stok']

        # Update produk
        cur.execute("""
            UPDATE produk SET nama_produk=%s, harga=%s, stok=%s WHERE id=%s
        """, (nama_produk, harga, stok, id))
        mysql.connection.commit()
        cur.close()

        flash('Produk berhasil diperbarui', 'success')
        return redirect(url_for('produk_list'))  # Ganti sesuai nama_produk fungsi daftar produk kamu

    cur.close()
    return render_template('edit_produk.html', produk=produk)

# ========== HAPUS PRODUK ==========
@app.route('/produk/delete/<int:id>', methods=['POST', 'GET'])
@login_required
@admin_required
def hapus_produk(id):
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    # Ambil nama file foto terlebih dahulu
    cur.execute("SELECT foto FROM produk WHERE id = %s", (id,))
    produk = cur.fetchone()

    if produk and produk['foto']:
        # Buat path lengkap menuju file
        foto_path = os.path.join(current_app.root_path, 'static', 'uploads', produk['foto'])
        if os.path.exists(foto_path):
            os.remove(foto_path)

    # Hapus data dari database
    cur.execute("DELETE FROM produk WHERE id = %s", (id,))
    mysql.connection.commit()
    cur.close()

    flash("Produk dan fotonya berhasil dihapus", "success")
    return redirect(url_for('produk_list'))

@app.route('/produk/hapus_all', methods=['POST'])
@login_required
@admin_required
def hapus_semua_produk():
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    # Ambil semua nama file foto sebelum data dihapus
    cur.execute("SELECT foto FROM produk")
    semua_foto = cur.fetchall()

    # Hapus file foto dari folder uploads
    for item in semua_foto:
        if item['foto']:
            foto_path = os.path.join(current_app.root_path, 'static', 'uploads', item['foto'])
            if os.path.exists(foto_path):
                os.remove(foto_path)

    # Hapus semua data dari tabel produk
    cur.execute("DELETE FROM produk")

    mysql.connection.commit()
    cur.close()

    flash('Semua produk dan fotonya berhasil dihapus', 'success')
    return redirect(url_for('produk_list'))

# ========== DETAIL PRODUK ==========
@app.route('/produk/<int:id>')
@login_required
def produk_detail(id):
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    cur.execute('SELECT * FROM produk WHERE id = %s', (id,))
    produk = cur.fetchone()
    cur.close()
    return render_template('produk_detail.html', produk=produk)

# ========== DETAIL PEmbelian ==========
@app.route('/pembelian')
@login_required
def lihat_pembelian():
    q = request.args.get('q', '')  # Pencarian nama pembelian
    start_date = request.args.get('start_date', '')  # Tanggal mulai
    end_date = request.args.get('end_date', '')  # Tanggal akhir

    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
    base_query = "SELECT * FROM pembelian WHERE 1=1"  # Query dasar untuk mengambil data
    params = []

    if q:
        base_query += " AND nama LIKE %s"
        params.append('%' + q + '%')  # Filter berdasarkan nama pembelian

    if start_date and end_date:
        try:
            # Mengubah tanggal mulai dan tanggal akhir menjadi objek datetime
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d')

            # Filter berdasarkan rentang tanggal
            base_query += " AND tgl_pembelian BETWEEN %s AND %s"
            params.extend([start_date_obj.date(), end_date_obj.date()])
        except:
            pass

    base_query += " ORDER BY tgl_pembelian DESC"  # Urutkan berdasarkan tanggal pembelian
    cur.execute(base_query, tuple(params))  # Eksekusi query dengan parameter
    pembelian_data = cur.fetchall()  # Ambil hasil query
    cur.close()

    return render_template('pembelian_list.html', pembelian=pembelian_data)

@app.route('/pembelian/export-excel')
@login_required
def export_pembelian_excel():
    start_date = request.args.get('start_date', '')  # Tanggal mulai
    end_date = request.args.get('end_date', '')  # Tanggal akhir
    q = request.args.get('q', '')  # Pencarian nama pembelian

    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    query = """
        SELECT 
            p.id,
            p.nama AS nama_pelanggan,
            p.role_pembuat AS dibuat_oleh,
            p.total AS total_harga,
            p.total_bayar,
            p.diskon,
            p.kembalian,
            p.sisa_point,
            p.nomor_pembayaran,
            p.tgl_pembelian,
            m.no_hp,
            m.status AS status_member,
            m.name AS nama_member,
            m.point AS member_point,
            m.join_date
        FROM pembelian p
        LEFT JOIN member m ON p.member_id = m.id
        WHERE 1=1
    """
    params = []

    # Filter berdasarkan nama
    if q:
        query += " AND p.nama LIKE %s"
        params.append('%' + q + '%')

    # Filter berdasarkan rentang tanggal (start_date dan end_date)
    if start_date and end_date:
        try:
            # Mengubah start_date dan end_date menjadi objek datetime
            start_date_obj = datetime.strptime(start_date, '%Y-%m-%d')
            end_date_obj = datetime.strptime(end_date, '%Y-%m-%d')

            # Filter berdasarkan rentang tanggal
            query += " AND p.tgl_pembelian BETWEEN %s AND %s"
            params.extend([start_date_obj.date(), end_date_obj.date()])
        except:
            pass

    query += " ORDER BY p.tgl_pembelian DESC"  # Urutkan berdasarkan tanggal pembelian
    cur.execute(query, tuple(params))  # Eksekusi query dengan parameter
    pembelian_list = cur.fetchall()

    # Ambil detail produk
    pembelian_produk_map = {}
    for pembelian in pembelian_list:
        cur.execute("""
            SELECT prod.nama_produk, prod.harga, lp.quantity, lp.subtotal
            FROM list_pembelian lp
            JOIN produk prod ON lp.produk_id = prod.id
            WHERE lp.pembelian_id = %s
        """, (pembelian['id'],))
        pembelian_produk_map[pembelian['id']] = cur.fetchall()

    cur.close()

    # Buat file Excel
    wb = Workbook()
    ws = wb.active
    ws.title = "Daftar Pembelian"

    headers = [
        'No', 'Nomor Pembayaran', 'Tanggal Pembelian', 'Nama Pelanggan', 'No HP',
        'Status Member', 'Nama Member', 'Sisa Point', 'Reward Point', 'Dibuat Oleh',
        'Total', 'Diskon', 'Bayar', 'Kembalian', 'Detail Produk'
    ]
    ws.append(headers)

    header_font = Font(bold=True)
    for col_num, column_title in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')

    persen = 0.01

    for idx, p in enumerate(pembelian_list, start=1):
        reward_point = round(float(p['total_harga']) * persen, 2)
        produk_list = pembelian_produk_map.get(p['id'], [])
        detail_produk = '; '.join([f"{prod['nama_produk']} (x{prod['quantity']} @ {format_rupiah(prod['harga'])})" for prod in produk_list])
        ws.append([
            idx,
            p['nomor_pembayaran'],
            p['tgl_pembelian'].strftime('%d-%m-%Y') if p['tgl_pembelian'] else '',
            p['nama_pelanggan'],
            p.get('no_hp') or '-',
            p.get('status_member') or '-',
            p.get('nama_member') or '-',
            int(p['sisa_point']) if p['sisa_point'] else 0,
            reward_point,
            p['dibuat_oleh'],
            format_rupiah(p['total_harga']),
            format_rupiah(p['diskon']),
            format_rupiah(p['total_bayar']),
            format_rupiah(p['kembalian']),
            detail_produk
        ])

    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    file_stream = BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    return send_file(
        file_stream,
        as_attachment=True,
        download_name='daftar_pembelian.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

@app.route('/pembelian/<int:pembelian_id>')
@login_required
def detail_pembelian(pembelian_id):
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    # Ambil data utama pembelian + member
    cur.execute("""
        SELECT p.id, p.nama AS nama_pelanggan, p.tgl_pembelian, p.role_pembuat, p.total,p.total_bayar, p.diskon, p.kembalian, p.sisa_point, p.nomor_pembayaran,
               m.no_hp, m.status AS member_status, m.join_date, m.point AS member_point, m.name AS member_name
        FROM pembelian p
        LEFT JOIN member m ON p.member_id = m.id
        WHERE p.id = %s
    """, (pembelian_id,))
    pembelian = cur.fetchone()

    if not pembelian:
        flash("Data pembelian tidak ditemukan.", "danger")
        return redirect(url_for('lihat_pembelian'))

    # Ambil daftar produk yang dibeli
    cur.execute("""
        SELECT prod.nama_produk AS nama_produk, prod.harga, lp.quantity, lp.subtotal, prod.foto
        FROM list_pembelian lp
        JOIN produk prod ON lp.produk_id = prod.id
        WHERE lp.pembelian_id = %s
    """, (pembelian_id,))
    list_produk = cur.fetchall()
    point = float(pembelian['total']) * persen

    cur.close()
    return render_template('detail_pembelian.html', pembelian=pembelian, list_produk=list_produk, reward_point=point)

@app.route('/pembelian/tambah', methods=['GET', 'POST'])
@login_required
def tambah_pembelian():
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    nomor_pembayaran = generate_nomor_pembayaran()

    if request.method == 'POST':
        user_id = session.get('user_id')
        role = session.get('role')
        email = session.get('email')

        # Ambil nama user
        cur.execute("SELECT name FROM users WHERE id = %s", (user_id,))
        user_data = cur.fetchone()
        if not user_data:
            flash('User tidak ditemukan', 'danger')
            return redirect(url_for('tambah_pembelian'))

        nama = user_data['name']
        tgl_pembelian = datetime.now()

        produk_ids = request.form.getlist('produk_id')
        quantities = request.form.getlist('quantity')
        name = request.form.get('name')
        no_hp = request.form.get('no_hp')
        name_member = request.form.get('name')
        is_member = request.form.get('is_member') == 'on'
        gunakan_point = request.form.get('gunakan_point') == 'on'

        total = 0
        total1 = 0
        detail_items = []

        member_data = None
        current_point = 0
        transaksi_pertama = False
        member_id = None

        total_bayar_input = request.form.get('total_bayar')
        try:
            total_bayar = float(total_bayar_input) if total_bayar_input else 0
        except ValueError:
            flash('Total Bayar tidak valid.', 'danger')
            return redirect(url_for('tambah_pembelian'))
        
        cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)
        cur.execute("SELECT * FROM users WHERE id = %s", (user_id,))
        user = cur.fetchone()

        for i in range(len(produk_ids)):
            produk_id = int(produk_ids[i])

            try:
                qty = int(quantities[i]) if quantities[i].strip() != '' else 0
            except ValueError:
                qty = 0

            if qty <= 0:
                continue

            cur.execute("SELECT stok, harga FROM produk WHERE id = %s", (produk_id,))
            produk1 = cur.fetchone()
            if not produk1:
                continue

            stok_tersedia1 = produk1['stok']
            harga1 = produk1['harga']

            if stok_tersedia1 < qty:
                flash(f'Stok produk ID {produk_id} tidak mencukupi', 'danger')
                return redirect(url_for('tambah_pembelian'))

            subtotal1 = harga1 * qty
            total1 += subtotal1

        diskon = 0
        
        point_digunakan = 0
        # Tambah point baru (10% dari total sebelum diskon)
        point_baru = (total1 * Decimal(persen))

        # Hitung sisa_point
        sisa_point = 0


        # Hitung total dari produk
        for i in range(len(produk_ids)):
            produk_id = int(produk_ids[i])

            try:
                qty = int(quantities[i]) if quantities[i].strip() != '' else 0
            except ValueError:
                qty = 0

            if qty <= 0:
                continue

            cur.execute("SELECT stok, harga FROM produk WHERE id = %s", (produk_id,))
            produk = cur.fetchone()
            if not produk:
                continue

            stok_tersedia = produk['stok']
            harga = produk['harga']

            if stok_tersedia < qty:
                flash(f'Stok produk ID {produk_id} tidak mencukupi', 'danger')
                return redirect(url_for('tambah_pembelian'))

            subtotal = harga * qty
            total += subtotal

            detail_items.append({
                'produk_id': produk_id,
                'quantity': qty,
                'price': harga,
                'subtotal': subtotal
            })

        if not detail_items:
            flash('Tidak ada produk yang dipilih untuk pembelian.', 'warning')
            return redirect(url_for('tambah_pembelian'))

        if is_member and gunakan_point and current_point > 0 and not transaksi_pertama:
            max_point_digunakan = total - 1  # total minimal harus tetap 1
            if max_point_digunakan <= 0:
                point_digunakan = 0
            else:
                point_digunakan = min(current_point, max_point_digunakan)

            diskon = point_digunakan
            total -= point_digunakan
            sisa_point = Decimal(current_point) - Decimal(point_digunakan) + Decimal(point_baru)
            kembalian = Decimal(str(total_bayar)) - Decimal(str(total))

        # Hitung kembalian
        kembalian = Decimal(str(total_bayar)) - Decimal(str(total))

        if total_bayar < total:
            flash(f'Total bayar ({total_bayar}) tidak mencukupi total pembelian ({total}).', 'danger')
            return redirect(url_for('tambah_pembelian'))

        if is_member and no_hp:
            cur.execute("SELECT * FROM member WHERE no_hp = %s", (no_hp,))
            member_data = cur.fetchone()
            if member_data:
                member_id = member_data['id']
                current_point = member_data['point']
                sisa_point = Decimal(current_point) + Decimal(point_baru)
                join_date = member_data.get('join_date')
                if join_date and join_date == datetime.today().date():
                    transaksi_pertama = False
            else:
                # Tambah member baru
                cur.execute("""
                    INSERT INTO member (name, no_hp, status, point, join_date)
                    VALUES (%s, %s, 'baru', 0, %s)
                """, (name_member,no_hp, datetime.now()))
                mysql.connection.commit()
                member_id = cur.lastrowid
                transaksi_pertama = True
                current_point = 0
                
        # Simpan ke tabel pembelian
        cur.execute("""
            INSERT INTO pembelian (user_id, nama, tgl_pembelian, total, role_pembuat, member_id, total_bayar, diskon, kembalian, sisa_point, nomor_pembayaran)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
        """, (user_id, nama, tgl_pembelian, total, role, member_id, total_bayar, diskon, kembalian, sisa_point, nomor_pembayaran))
        pembelian_id = cur.lastrowid

        # Simpan ke list_pembelian dan update stok
        for item in detail_items:
            cur.execute("""
                INSERT INTO list_pembelian (pembelian_id, produk_id, quantity, price, subtotal)
                VALUES (%s, %s, %s, %s, %s)
            """, (
                pembelian_id,
                item['produk_id'],
                item['quantity'],
                item['price'],
                item['subtotal']
            ))
            cur.execute("""
                UPDATE produk
                SET stok = stok - %s
                WHERE id = %s
            """, (item['quantity'], item['produk_id']))

        if is_member:
            point_sisa = Decimal(current_point) - Decimal(point_digunakan) + Decimal(point_baru)
            cur.execute("""
                UPDATE member
                SET point = %s
                WHERE id = %s
            """, (point_sisa, member_id))

        mysql.connection.commit()
        cur.close()
        flash(f'Pembelian berhasil! Point digunakan: {point_digunakan}, Point didapat: {round(point_baru, 2)}', 'success')
        return redirect(url_for('lihat_pembelian'))

    # GET method: tampilkan produk
    cur.execute("SELECT * FROM produk")
    produk = cur.fetchall()
    cur.close()

    return render_template('pembelian_form.html', produk=produk)

@app.route('/pembelian/hapus/<int:pembelian_id>', methods=['POST'])
@login_required
def hapus_pembelian(pembelian_id):
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    # Cek apakah user ingin mengembalikan stok
    kembalikan_stok = request.form.get('kembalikan_stok') == 'yes'

    if kembalikan_stok:
        # Ambil list pembelian untuk pembelian ini
        cur.execute("SELECT produk_id, quantity FROM list_pembelian WHERE pembelian_id = %s", (pembelian_id,))
        items = cur.fetchall()

        # Kembalikan stok produk
        for item in items:
            cur.execute("""
                UPDATE produk
                SET stok = stok + %s
                WHERE id = %s
            """, (item['quantity'], item['produk_id']))

    # Hapus list pembelian dan pembeliannya
    cur.execute("DELETE FROM list_pembelian WHERE pembelian_id = %s", (pembelian_id,))
    cur.execute("DELETE FROM pembelian WHERE id = %s", (pembelian_id,))
    mysql.connection.commit()
    cur.close()

    flash('Pembelian berhasil dihapus' + (' dan stok dikembalikan.' if kembalikan_stok else '.'), 'success')
    return redirect(url_for('lihat_pembelian'))

@app.route('/pembelian/hapus_all', methods=['POST'])
@login_required
def hapus_semua_pembelian():
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    kembalikan_stok = request.form.get('kembalikan_stok') == 'yes'

    if kembalikan_stok:
        # Ambil semua pembelian dan listnya
        cur.execute("SELECT produk_id, quantity FROM list_pembelian")
        items = cur.fetchall()

        for item in items:
            cur.execute("""
                UPDATE produk
                SET stok = stok + %s
                WHERE id = %s
            """, (item['quantity'], item['produk_id']))

    # Hapus semua list_pembelian dan pembelian
    cur.execute("DELETE FROM list_pembelian")
    cur.execute("DELETE FROM pembelian")

    mysql.connection.commit()
    cur.close()

    flash('Semua pembelian berhasil dihapus' + (' dan stok dikembalikan.' if kembalikan_stok else '.'), 'success')
    return redirect(url_for('lihat_pembelian'))

from pathlib import Path

@app.route('/pembelian/<int:pembelian_id>/pdf')
@login_required
def unduh_pembelian_pdf(pembelian_id):
    cur = mysql.connection.cursor(MySQLdb.cursors.DictCursor)

    # Ambil data pembelian
    cur.execute("""
        SELECT p.id, p.nama AS nama_pelanggan, p.tgl_pembelian, p.role_pembuat, p.total, p.total_bayar, p.diskon, p.kembalian, p.sisa_point, p.nomor_pembayaran,
               m.no_hp, m.status AS member_status, m.join_date, m.point AS member_point, m.name AS member_name
        FROM pembelian p
        LEFT JOIN member m ON p.member_id = m.id
        WHERE p.id = %s
    """, (pembelian_id,))
    pembelian = cur.fetchone()

    if not pembelian:
        flash("Data pembelian tidak ditemukan.", "danger")
        return redirect(url_for('lihat_pembelian'))

    # Ambil daftar produk
    cur.execute("""
        SELECT prod.nama_produk AS nama_produk, prod.harga, lp.quantity, lp.subtotal
        FROM list_pembelian lp
        JOIN produk prod ON lp.produk_id = prod.id
        WHERE lp.pembelian_id = %s
    """, (pembelian_id,))
    list_produk = cur.fetchall()

    # Hitung reward point
    point = float(pembelian['total']) * persen

    # Buat PDF
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []

    styles = getSampleStyleSheet()
    normal = styles["Normal"]
    bold = styles["Heading4"]
    title = styles["Title"]

    # Judul utama
    elements.append(Paragraph("INVOICE PENJUALAN", title))
    elements.append(Spacer(1, 12))

    # Informasi di bagian atas
    info_atas = [
        ["Nomor Pembayaran", pembelian["nomor_pembayaran"]],
        ["Tanggal Pembelian", pembelian["tgl_pembelian"]],
        ["Nama Pelanggan", pembelian["nama_pelanggan"]],
        ["No Telephon", pembelian["no_hp"]],
    ]
    info_kanan = [
        ["Status Member", pembelian["member_status"]],
        ["Sisa Point", int(pembelian["sisa_point"])],
        ["Reward Point", round(point, 2)],
        ["Dibuat oleh", pembelian["role_pembuat"]],
    ]
    gabung_info = []
    for kiri, kanan in zip(info_atas, info_kanan):
        gabung_info.append([f"{kiri[0]}: {kiri[1]}", f"{kanan[0]}: {kanan[1]}"])

    table_info = Table(gabung_info, colWidths=[8 * cm, 8 * cm])
    table_info.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
    ]))
    elements.append(table_info)
    elements.append(Spacer(1, 20))

    # Tabel Produk
    # Gabungkan Tabel Produk + Ringkasan Pembayaran
    elements.append(Paragraph("Daftar Produk", bold))

    data_produk = [["No", "Nama Produk", "Harga", "Jumlah", "Subtotal"]]
    for idx, item in enumerate(list_produk, start=1):
        data_produk.append([
            idx,
            item["nama_produk"],
            format_rupiah(item['harga']),
            item["quantity"],
            format_rupiah(item['subtotal']),
        ])

    table_produk = Table(data_produk, colWidths=[1.5*cm, 5.5*cm, 3.5*cm, 2*cm, 3.5*cm])
    table_produk.setStyle(TableStyle([
        ('GRID', (0, 0), (-1, -1), 0.4, colors.black),
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ALIGN', (2, 1), (-1, -1), 'RIGHT'),  # Harga & subtotal rata kanan
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # "No" kolom rata tengah
        ('ALIGN', (3, 1), (3, -1), 'CENTER'),  # "Jumlah" kolom rata tengah
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),  # Header tabel rata tengah
    ]))
    elements.append(table_produk)
    elements.append(Spacer(1, 5))  # Spacer kecil agar lebih dekat dengan subtotal

    # Ringkasan Pembayaran sebagai tabel rapi
    data_ringkasan = [
        ["", "", "", "Total", format_rupiah(pembelian['total'])],
        ["", "", "", "Diskon", "-" + format_rupiah(pembelian['diskon'])],
        ["", "", "", "Bayar", format_rupiah(pembelian['total_bayar'])],
        ["", "", "", "Kembalian", format_rupiah(pembelian['kembalian'])],
    ]

    table_ringkasan = Table(data_ringkasan, colWidths=[1.5*cm, 5.5*cm, 3.5*cm, 2*cm, 3.5*cm])
    table_ringkasan.setStyle(TableStyle([
        ('ALIGN', (3, 0), (3, -1), 'RIGHT'),   # Label: Total, Diskon, etc
        ('ALIGN', (4, 0), (4, -1), 'RIGHT'),   # Value: Rp...
        ('FONTNAME', (3, 0), (4, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (3, 0), (4, -1), 10),
        ('TOPPADDING', (3, 0), (4, -1), 2),
        ('BOTTOMPADDING', (3, 0), (4, -1), 2),
    ]))
    elements.append(table_ringkasan)

    doc.build(elements)

    pdf = buffer.getvalue()
    buffer.close()

    response = make_response(pdf)
    response.headers["Content-Type"] = "application/pdf"
    response.headers["Content-Disposition"] = f"attachment; filename=pembelian_{pembelian_id}.pdf"
    return response

# ========== LOGOUT ==========
@app.route('/logout')
def logout():
    session.clear()
    flash('Berhasil logout.', 'info')
    return redirect(url_for('login'))

# ========== JALANKAN APLIKASI ==========
# if __name__ == '__main__':
#     app.run(debug=True)